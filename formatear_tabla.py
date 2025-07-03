import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Ruta de la carpeta donde están los archivos Excel
ruta_carpeta = r"C:\Users\David Yail\Desktop\Nueva carpeta"

# Procesar cada archivo Excel en la carpeta
for archivo in os.listdir(ruta_carpeta):
    if archivo.endswith('.xlsx') and 'LP_Profar' in archivo:
        ruta_completa = os.path.join(ruta_carpeta, archivo)
        
        try:
            # Cargar el archivo Excel
            wb = load_workbook(ruta_completa)
            ws = wb['LP_PROFAR']
            
            # Encontrar la columna de fecha
            fecha_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'Fecha':
                    fecha_col = col
                    break
            
            # Aplicar formato de fecha a toda la columna
            if fecha_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=fecha_col)
                    cell.number_format = 'dd-mm-yyyy'
            
            # Crear la tabla
            tabla = Table(displayName="LP_PROFAR", ref=f"A1:{chr(64 + ws.max_column)}{ws.max_row}")
            
            # Aplicar estilo a la tabla
            estilo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tabla.tableStyleInfo = estilo
            
            # Agregar la tabla a la hoja
            ws.add_table(tabla)
            
            # Guardar los cambios
            wb.save(ruta_completa)
            print(f"Archivo formateado exitosamente: {archivo}")
                
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {str(e)}")

print("Proceso completado.") 