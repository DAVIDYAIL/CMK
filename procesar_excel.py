import pandas as pd
import os
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import numbers

# Ruta de la carpeta donde están los archivos Excel
ruta_carpeta = r"C:\Users\David Yail\Desktop\Nueva carpeta"

# Función para extraer la fecha del nombre del archivo
def extraer_fecha(nombre_archivo):
    # Busca el patrón de fecha en el nombre del archivo
    patron = r'(\d{2}-\d{2}-\d{4})'
    coincidencia = re.search(patron, nombre_archivo)
    if coincidencia:
        fecha_str = coincidencia.group(1)
        # Convertir la fecha a formato datetime
        return datetime.strptime(fecha_str, '%d-%m-%Y')
    return None

# Procesar cada archivo Excel en la carpeta
for archivo in os.listdir(ruta_carpeta):
    if archivo.endswith('.xlsx') and 'LP_Profar' in archivo:
        ruta_completa = os.path.join(ruta_carpeta, archivo)
        
        try:
            # Leer el archivo Excel
            df = pd.read_excel(ruta_completa, sheet_name='LP_PROFAR')
            
            # Extraer la fecha del nombre del archivo
            fecha = extraer_fecha(archivo)
            
            if fecha is not None:
                # Agregar la columna de fecha
                df['Fecha'] = fecha.date()  # Convertir a solo fecha sin hora
                
                # Guardar el archivo modificado
                df.to_excel(ruta_completa, sheet_name='LP_PROFAR', index=False)
                
                # Aplicar formato de fecha a la columna
                wb = load_workbook(ruta_completa)
                ws = wb['LP_PROFAR']
                
                # Encontrar la columna de fecha
                fecha_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == 'Fecha':
                        fecha_col = col
                        break
                
                if fecha_col:
                    # Aplicar formato de fecha a toda la columna
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=fecha_col)
                        cell.number_format = 'dd-mm-yyyy'
                
                wb.save(ruta_completa)
                print(f"Archivo procesado exitosamente: {archivo}")
            else:
                print(f"No se pudo extraer la fecha del archivo: {archivo}")
                
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {str(e)}")

print("Proceso completado.") 